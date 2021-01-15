"""ABC
"""
import os
import openpyxl
import xlsxwriter
from firebase import firebase


class Filemanager:
    """Edite des fichiers de types TXT, XML ou MD
    """
    def __init__(self, quote_lst):
        self.quote_lst = quote_lst

    def edit_txt_file(self):
        """Edite un fichier TXT avec les tags présents dans la liste
        de quote en paramètres
        (1 par ligne) dans le dossier Result de la racine projet
        Parameter:
        quote_lst (Array of Quotes): liste de citations
        """

        lines = []
        tags_lst = []
        for quote in self.quote_lst:
            for tag in quote.tags:
                if tag not in tags_lst:
                    tags_lst.append(tag+'\n')
        mode = "r" if os.path.isfile("Result/tags.txt") else "x"
        with open("Result/tags.txt", mode) as file:
            lines = [] \
                    if mode == "x" or os.stat("Result/tags.txt").st_size == 0 \
                    else file.readlines()
            for tag in tags_lst:
                if tag not in lines:
                    file.write(tag)
            file.close()

    def edit_md_file(self):
        """Edite un fichier MD avec les citations, les
        auteurs et les tags présents dans la liste
        de quote en paramètres ans le dossier Result
        de la racine projet.

        Parameter:
        quote_lst (Array of Quotes): liste de citations
        """

        mode = "w" if os.path.isfile("Result/quotes.md") else "x"
        with open("Result/quotes.md", mode) as file:
            file.write("Citations | Auteur | Tags \n")
            file.write("---|---|---\n")
            for quote in self.quote_lst:
                # print(quote.text)
                
                print(', '.join(quote.tags)+'\n'
                      if len(quote.tags) > 0
                      else '')
                file.write(quote.text + " | "
                           + quote.author.name + " | "
                           + (', '.join(quote.tags)+'\n'
                              if len(quote.tags) > 0
                              else '\n'))

    def edit_xlsx_file(self):
        """Edite un fichier XLSX avec le nom de l'auteur,
        sa date de naissance et sa biographie
        dans le dossier Result de la racine projet
        Parameter:
        quote_lst (Array of Quotes): liste de citations
        """

        try:
            file = openpyxl.load_workbook('Result/Autheurs.xlsx')
        except IOError:
            workbook = xlsxwriter.Workbook('Result/Autheurs.xlsx')
            workbook.close()
            file = openpyxl.load_workbook('Result/Autheurs.xlsx')
        worksheet = file.active
        worksheet.cell(row=1, column=1).value = "Nom"
        worksheet.cell(row=1, column=2).value = "Date de naissance"
        worksheet.cell(row=1, column=3).value = "Description"
        i = 2
        for quote in self.quote_lst:
          #  Insert Bdd
            self.firebase = firebase.FirebaseApplication('https://pythoncnam-default-rtdb.europe-west1.firebasedatabase.app/', None)
            self.data =  {'Name': quote.author.name,
                            'Naissance': quote.author.birthdate,
                            'Description' : quote.author.description  }
            self.result = self.firebase.post('pythoncnam/Auteur',self.data) 

            worksheet.cell(row=i, column=1).value = quote.author.name
            worksheet.cell(row=i, column=2).value = quote.author.birthdate
            worksheet.cell(row=i, column=3).value = quote.author.description
            i += 1
        file.save('Result/Autheurs.xlsx')
